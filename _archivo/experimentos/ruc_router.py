# ruc_router.py
# Agregar a tu FastAPI app:
#
#   from ruc_router import router as ruc_router
#   app.include_router(ruc_router)
#
# Instalar: pip install thefuzz python-levenshtein openpyxl

import io
import re
import time
import uuid
import threading
import requests
import pandas as pd
from thefuzz import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from fastapi import APIRouter, File, Form, UploadFile, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from typing import List

router = APIRouter(prefix="/api/ruc", tags=["RUC"])

# ── Estado de jobs en memoria ─────────────────────────────────
# Estructura: { job_id: { status, progress, total, done, error, result_bytes } }
_jobs: dict = {}
_jobs_lock = threading.Lock()

# ── Cache de búsquedas API ────────────────────────────────────
_cache: dict = {}
_cache_lock = threading.Lock()

# ── Configuración ─────────────────────────────────────────────
TURUC_URL      = "https://turuc.com.py/api/contribuyente/search"
UMBRAL_AUTO    = 82
UMBRAL_REVISAR = 60
DELAY          = 0.35   # segundos entre llamadas
MAX_PAGINAS    = 2
JOB_TTL        = 3600   # segundos hasta limpiar job de memoria

SESSION = requests.Session()
SESSION.headers.update({"User-Agent": "JARVIS-RUC/1.0 Paraguay"})

SUFIJOS = [r"\bS\.A\.C\b", r"\bS\.A\b", r"\bS\.R\.L\.\b", r"\bS\.R\.L\b",
           r"\bSRL\b", r"\bSAC\b", r"\bSAS\b", r"\bSA\b", r"\bLTDA\b",
           r"\bCIA\b", r"\bCÍA\b", r"\bE\.I\.R\.L\b", r"\bSPA\b"]
STOPS = {"DE","DEL","LA","LAS","LOS","EL","Y","E","A","EN","CON","POR","PARA","SU","SUS"}

FILLS = {
    "AUTO":          PatternFill("solid", fgColor="C6EFCE"),
    "REVISAR":       PatternFill("solid", fgColor="FFEB9C"),
    "NO_MATCH":      PatternFill("solid", fgColor="FFC7CE"),
    "NO_ENCONTRADO": PatternFill("solid", fgColor="FFC7CE"),
    "SIN_DATO":      PatternFill("solid", fgColor="D9D9D9"),
}


# ── Helpers ───────────────────────────────────────────────────

def _normalizar(texto: str) -> str:
    if not isinstance(texto, str): return ""
    t = texto.upper().strip()
    for s in SUFIJOS: t = re.sub(s, "", t)
    return re.sub(r"\s+", " ", re.sub(r"[^\w\s]", " ", t)).strip()

def _query(razon: str) -> str:
    palabras = [p for p in _normalizar(razon).split() if len(p) > 2 and p not in STOPS]
    return " ".join(palabras[:3]) or razon[:20].strip()

def _similitud(a: str, b: str) -> float:
    return max(fuzz.token_set_ratio(a, b), fuzz.partial_ratio(a, b) * 0.85, fuzz.ratio(a, b) * 0.8)

def _buscar_api(termino: str) -> list:
    with _cache_lock:
        if termino in _cache: return _cache[termino]
    candidatos = []
    for page in range(MAX_PAGINAS):
        try:
            r = SESSION.get(TURUC_URL, params={"search": termino, "page": page}, timeout=12)
            if r.status_code != 200: break
            data = r.json()
            items = data.get("data", {}).get("contribuyentes", [])
            if not items: break
            candidatos.extend(items)
            if page + 1 >= data.get("data", {}).get("paginas", 1): break
            time.sleep(0.15)
        except Exception:
            break
    with _cache_lock:
        _cache[termino] = candidatos
    return candidatos

def _match_uno(razon: str) -> dict:
    vacio = {"ruc": "", "razon_api": "", "similitud": 0, "estado": "", "resultado": "SIN_DATO"}
    if not isinstance(razon, str) or not razon.strip():
        return vacio
    q = _query(razon)
    candidatos = _buscar_api(q) if len(q) >= 3 else []
    if not candidatos:
        for p in sorted(_normalizar(razon).split(), key=len, reverse=True)[:2]:
            if len(p) >= 4:
                candidatos = _buscar_api(p)
                if candidatos: break
    if not candidatos:
        return {**vacio, "resultado": "NO_ENCONTRADO"}
    norm = _normalizar(razon)
    mejor, sc_max = None, 0
    for c in candidatos:
        sc = _similitud(norm, _normalizar(c.get("razonSocial", "")))
        if sc > sc_max: sc_max, mejor = sc, c
    if not mejor: return {**vacio, "resultado": "NO_ENCONTRADO"}
    ruc = mejor.get("ruc", "")
    resultado = ("AUTO" if sc_max >= UMBRAL_AUTO else "REVISAR" if sc_max >= UMBRAL_REVISAR else "NO_MATCH")
    return {
        "ruc":       ruc if resultado != "NO_MATCH" else "",
        "razon_api": mejor.get("razonSocial", ""),
        "similitud": round(sc_max, 1),
        "estado":    mejor.get("estado", ""),
        "resultado": resultado,
    }

def _encontrar_col(df: pd.DataFrame, nombre: str) -> str | None:
    n = nombre.lower().strip()
    for col in df.columns:
        if str(col).lower().strip() == n: return col
    return None

def _df_to_xlsx(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    for cell in ws[1]: cell.font = Font(bold=True, name="Arial")
    col_idx = next((i for i, c in enumerate(ws[1], 1) if c.value == "Resultado_Busqueda"), None)
    if col_idx:
        for row in ws.iter_rows(min_row=2):
            fill = FILLS.get(row[col_idx - 1].value)
            if fill:
                for cell in row: cell.fill = fill
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 3, 55
        )
    ws.freeze_panes = "A2"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Tarea en background ───────────────────────────────────────

def _procesar_job(job_id: str, df: pd.DataFrame, col_rs: str):
    """Corre en un thread separado. Actualiza _jobs[job_id] a medida que avanza."""
    total = len(df)
    ya_vistos: dict = {}
    stats = {"AUTO": 0, "REVISAR": 0, "NO_MATCH": 0, "NO_ENCONTRADO": 0, "SIN_DATO": 0}

    for col in ["RUC_Encontrado", "RazonSocial_API", "Estado_Contribuyente", "Resultado_Busqueda"]:
        df[col] = ""
    df["Similitud_%"] = 0.0

    try:
        for idx, row in df.iterrows():
            rs = str(row[col_rs]).strip() if pd.notna(row[col_rs]) else ""
            key = rs.upper().strip()
            if key not in ya_vistos:
                ya_vistos[key] = _match_uno(rs)
                time.sleep(DELAY)
            res = ya_vistos[key]
            df.at[idx, "RUC_Encontrado"]       = res["ruc"]
            df.at[idx, "RazonSocial_API"]       = res["razon_api"]
            df.at[idx, "Similitud_%"]           = res["similitud"]
            df.at[idx, "Estado_Contribuyente"]  = res["estado"]
            df.at[idx, "Resultado_Busqueda"]    = res["resultado"]
            stats[res["resultado"]] = stats.get(res["resultado"], 0) + 1

            # Actualizar progreso cada 10 filas
            if idx % 10 == 0:
                with _jobs_lock:
                    _jobs[job_id]["procesadas"] = idx + 1
                    _jobs[job_id]["progress"]   = round((idx + 1) / total * 100, 1)

        # Finalizar
        xlsx_bytes = _df_to_xlsx(df)
        with _jobs_lock:
            _jobs[job_id].update({
                "status":    "done",
                "progress":  100.0,
                "procesadas": total,
                "stats":     stats,
                "result":    xlsx_bytes,
                "finished_at": time.time(),
            })
    except Exception as e:
        with _jobs_lock:
            _jobs[job_id].update({"status": "error", "error": str(e)})


# ── Endpoints ─────────────────────────────────────────────────

@router.post("/enriquecer")
async def enriquecer(
    files: List[UploadFile] = File(...),
    col_razon_social: str   = Form(...),
):
    """
    Recibe uno o más Excel, lanza el procesamiento en background.
    Retorna inmediatamente con { job_id }.
    """
    if not files:
        raise HTTPException(400, "No se recibieron archivos")

    dfs = []
    errores = []
    for f in files:
        if not f.filename.endswith((".xlsx", ".xls")):
            errores.append(f"{f.filename}: formato no soportado")
            continue
        try:
            contents = await f.read()
            df = pd.read_excel(io.BytesIO(contents), dtype=str)
            df["__archivo_origen"] = f.filename
            dfs.append(df)
        except Exception as e:
            errores.append(f"{f.filename}: {e}")

    if not dfs:
        raise HTTPException(400, {"error": "No se pudo leer ningún archivo", "detalle": errores})

    df_total = pd.concat(dfs, ignore_index=True)

    col_rs = _encontrar_col(df_total, col_razon_social)
    if col_rs is None:
        disponibles = [c for c in df_total.columns if not c.startswith("__")]
        raise HTTPException(400, {
            "error": f"Columna '{col_razon_social}' no encontrada",
            "columnas_disponibles": disponibles,
        })

    job_id = str(uuid.uuid4())
    with _jobs_lock:
        _jobs[job_id] = {
            "status":    "processing",
            "progress":  0.0,
            "procesadas": 0,
            "total":     len(df_total),
            "archivos":  len(dfs),
            "stats":     {},
            "result":    None,
            "error":     None,
            "created_at": time.time(),
        }

    t = threading.Thread(target=_procesar_job, args=(job_id, df_total, col_rs), daemon=True)
    t.start()

    return JSONResponse({"job_id": job_id, "total": len(df_total), "archivos": len(dfs)})


@router.get("/status/{job_id}")
def status(job_id: str):
    """
    Polling endpoint. El frontend llama esto cada 2 segundos.
    Retorna: { status, progress, procesadas, total, stats?, error? }
    """
    with _jobs_lock:
        job = _jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job no encontrado")

    resp = {
        "status":    job["status"],       # processing | done | error
        "progress":  job["progress"],     # 0.0 - 100.0
        "procesadas": job["procesadas"],
        "total":     job["total"],
    }
    if job["status"] == "done":
        resp["stats"] = job["stats"]
    if job["status"] == "error":
        resp["error"] = job["error"]
    return JSONResponse(resp)


@router.get("/download/{job_id}")
def download(job_id: str):
    """
    Descarga el Excel resultado. Llamar solo cuando status == 'done'.
    Limpia el job de memoria después de 1 hora (TTL).
    """
    with _jobs_lock:
        job = _jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job no encontrado")
    if job["status"] != "done":
        raise HTTPException(400, f"Job no terminado aún (status: {job['status']})")
    if not job["result"]:
        raise HTTPException(500, "Resultado no disponible")

    return StreamingResponse(
        io.BytesIO(job["result"]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=resultado_ruc.xlsx"},
    )


@router.post("/preview")
async def preview(file: UploadFile = File(...)):
    """
    Devuelve columnas y 5 primeras filas de un Excel.
    Usar antes de llamar a /enriquecer para que el usuario elija la columna.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Solo .xlsx o .xls")
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents), dtype=str, nrows=5)
        return JSONResponse({
            "columnas": list(df.columns),
            "preview":  df.fillna("").to_dict(orient="records"),
            "filename": file.filename,
        })
    except Exception as e:
        raise HTTPException(500, str(e))


@router.get("/buscar")
def buscar_uno(q: str):
    """
    Lookup rápido de un nombre. Útil para el chat en tiempo real.
    GET /api/ruc/buscar?q=COCA+COLA
    """
    if len(q.strip()) < 3:
        raise HTTPException(400, "Mínimo 3 caracteres")
    return JSONResponse(_match_uno(q.strip()))


@router.delete("/job/{job_id}")
def limpiar_job(job_id: str):
    """Limpieza manual de un job. Llamar después de descargar el resultado."""
    with _jobs_lock:
        if job_id in _jobs:
            del _jobs[job_id]
    return JSONResponse({"ok": True})
