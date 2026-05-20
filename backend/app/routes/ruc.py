import io
import re
import time
import uuid
import json
import zipfile
import unicodedata
import threading
import requests
import pandas as pd
from thefuzz import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from fastapi import APIRouter, File, Form, UploadFile, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from typing import List, Optional

router = APIRouter(prefix="/api/ruc", tags=["RUC"])

# ── Estado de jobs en memoria ─────────────────────────────────
_jobs: dict = {}
_jobs_lock = threading.Lock()

# ── Cache de búsquedas API ────────────────────────────────────
_cache: dict = {}
_cache_lock = threading.Lock()

# ── Configuración ─────────────────────────────────────────────
TURUC_URL      = "https://turuc.com.py/api/contribuyente/search"
UMBRAL_AUTO    = 82
UMBRAL_REVISAR = 60
DELAY          = 0.35
MAX_PAGINAS    = 2

SESSION = requests.Session()
SESSION.headers.update({"User-Agent": "JARVIS-RUC/1.0 Paraguay"})

SUFIJOS = [r"\bS\.A\.C\b", r"\bS\.A\b", r"\bS\.R\.L\.\b", r"\bS\.R\.L\b",
           r"\bSRL\b", r"\bSAC\b", r"\bSAS\b", r"\bSA\b", r"\bLTDA\b",
           r"\bCIA\b", r"\bCÍA\b", r"\bE\.I\.R\.L\b", r"\bSPA\b"]
STOPS = {"DE","DEL","LA","LAS","LOS","EL","Y","E","A","EN","CON","POR","PARA","SU","SUS"}

FILLS = {
    "AUTO":           PatternFill("solid", fgColor="C6EFCE"),
    "REVISAR":        PatternFill("solid", fgColor="FFEB9C"),
    "NO_MATCH":       PatternFill("solid", fgColor="FFC7CE"),
    "NO_ENCONTRADO":  PatternFill("solid", fgColor="FFC7CE"),
    "SIN_DATO":       PatternFill("solid", fgColor="D9D9D9"),
    "RUC_COMPLETADO": PatternFill("solid", fgColor="C6EFCE"),
}


# ── Helpers para búsqueda por nombre ─────────────────────────

def _normalizar(texto: str) -> str:
    """Normaliza texto: elimina tildes, sufijos legales y puntuación."""
    if not isinstance(texto, str):
        return ""
    # Quitar tildes / diacríticos
    t = unicodedata.normalize("NFD", texto)
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    t = t.upper().strip()
    for s in SUFIJOS:
        t = re.sub(s, "", t)
    return re.sub(r"\s+", " ", re.sub(r"[^\w\s]", " ", t)).strip()

def _similitud(a: str, b: str) -> float:
    return max(
        fuzz.token_set_ratio(a, b),
        fuzz.partial_ratio(a, b) * 0.85,
        fuzz.ratio(a, b) * 0.8,
    )

def _buscar_api(termino: str) -> list:
    with _cache_lock:
        if termino in _cache:
            return _cache[termino]
    candidatos = []
    for page in range(MAX_PAGINAS):
        try:
            r = SESSION.get(TURUC_URL, params={"search": termino, "page": page}, timeout=12)
            if r.status_code != 200:
                break
            data = r.json()
            items = data.get("data", {}).get("contribuyentes", [])
            if not items:
                break
            candidatos.extend(items)
            if page + 1 >= data.get("data", {}).get("paginas", 1):
                break
            time.sleep(0.15)
        except Exception:
            break
    with _cache_lock:
        _cache[termino] = candidatos
    return candidatos

def _match_uno(razon: str) -> dict:
    """
    Busca la empresa en turuc.com.py usando múltiples estrategias de query
    para maximizar la probabilidad de encontrar el RUC correcto.
    """
    vacio = {"ruc": "", "razon_api": "", "similitud": 0, "estado": "", "resultado": "SIN_DATO"}
    if not isinstance(razon, str) or not razon.strip():
        return vacio

    norm = _normalizar(razon)
    palabras = [p for p in norm.split() if len(p) > 2 and p not in STOPS]
    palabras_largas = [p for p in palabras if len(p) >= 4]

    # Acumula candidatos únicos por RUC (evita duplicados entre búsquedas)
    pool: dict[str, dict] = {}

    def _agregar(termino: str):
        if len(termino) < 3:
            return
        for c in _buscar_api(termino):
            ruc = c.get("ruc", "")
            if ruc and ruc not in pool:
                pool[ruc] = c

    # Estrategia 1: primeras 3 palabras significativas
    q1 = " ".join(palabras[:3]) or razon[:20].strip()
    _agregar(q1)

    # Estrategia 2: si no hay resultados, palabras largas individualmente
    if not pool:
        for p in sorted(palabras_largas, key=len, reverse=True)[:3]:
            _agregar(p)
            if pool:
                break

    if not pool:
        return {**vacio, "resultado": "NO_ENCONTRADO"}

    # Evaluar similitud con el pool actual
    mejor, sc_max = None, 0
    for c in pool.values():
        sc = _similitud(norm, _normalizar(c.get("razonSocial", "")))
        if sc > sc_max:
            sc_max, mejor = sc, c

    # Estrategia 3: si no llegamos a AUTO, probar queries alternativos
    if sc_max < UMBRAL_AUTO and palabras_largas:
        queries_extra = []
        # Par de palabras más largas
        if len(palabras_largas) >= 2:
            queries_extra.append(f"{palabras_largas[0]} {palabras_largas[1]}")
        # Segunda palabra sola (para casos tipo "BANCO CONTINENTAL" → "CONTINENTAL")
        if len(palabras_largas) >= 2:
            queries_extra.append(palabras_largas[1])
        # Palabra más larga sola
        queries_extra.append(max(palabras_largas, key=len))

        antes = len(pool)
        for q_extra in queries_extra:
            if q_extra != q1:
                _agregar(q_extra)

        # Re-evaluar solo si llegaron nuevos candidatos
        if len(pool) > antes:
            for c in pool.values():
                sc = _similitud(norm, _normalizar(c.get("razonSocial", "")))
                if sc > sc_max:
                    sc_max, mejor = sc, c

    if not mejor:
        return {**vacio, "resultado": "NO_ENCONTRADO"}

    resultado = "AUTO" if sc_max >= UMBRAL_AUTO else "REVISAR" if sc_max >= UMBRAL_REVISAR else "NO_MATCH"
    return {
        "ruc":       mejor.get("ruc", "") if resultado != "NO_MATCH" else "",
        "razon_api": mejor.get("razonSocial", ""),
        "similitud": round(sc_max, 1),
        "estado":    mejor.get("estado", ""),
        "resultado": resultado,
    }

def _encontrar_col(df: pd.DataFrame, nombre: str) -> str | None:
    n = nombre.lower().strip()
    for col in df.columns:
        if str(col).lower().strip() == n:
            return col
    return None

def _df_to_xlsx(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    for cell in ws[1]:
        cell.font = Font(bold=True, name="Arial")
    col_idx = next((i for i, c in enumerate(ws[1], 1) if c.value == "Resultado_RUC"), None)
    if col_idx:
        for row in ws.iter_rows(min_row=2):
            fill = FILLS.get(row[col_idx - 1].value)
            if fill:
                for cell in row:
                    cell.fill = fill
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 3, 55
        )
    ws.freeze_panes = "A2"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Tarea en background ───────────────────────────────────────

def _procesar_job(job_id: str, archivos_data: list):
    total_filas = sum(len(a["df"]) for a in archivos_data)
    filas_procesadas = 0

    try:
        for i, arch in enumerate(archivos_data):
            df = arch["df"].copy()
            fuente = arch["fuente_tipo"]
            col_rs = arch.get("col_razon")
            col_ruc = arch.get("col_ruc")

            stats = {"AUTO": 0, "REVISAR": 0, "NO_MATCH": 0,
                     "NO_ENCONTRADO": 0, "SIN_DATO": 0, "RUC_COMPLETADO": 0}
            no_encontrados = []

            df["RUC_Encontrado"] = ""
            df["Resultado_RUC"] = ""

            if True:
                df["RazonSocial_API"] = ""
                df["Similitud_%"] = 0.0
                df["Estado_Contribuyente"] = ""
                ya_vistos: dict = {}

                for idx, row in df.iterrows():
                    rs = str(row[col_rs]).strip() if col_rs and pd.notna(row.get(col_rs)) else ""
                    key = rs.upper().strip()
                    if key not in ya_vistos:
                        ya_vistos[key] = _match_uno(rs)
                        time.sleep(DELAY)
                    res = ya_vistos[key]
                    df.at[idx, "RUC_Encontrado"] = res["ruc"]
                    df.at[idx, "RazonSocial_API"] = res.get("razon_api", "")
                    df.at[idx, "Similitud_%"] = res.get("similitud", 0)
                    df.at[idx, "Estado_Contribuyente"] = res.get("estado", "")
                    df.at[idx, "Resultado_RUC"] = res["resultado"]
                    stats[res["resultado"]] = stats.get(res["resultado"], 0) + 1

                    if res["resultado"] in ("REVISAR", "NO_MATCH", "NO_ENCONTRADO") and rs:
                        no_encontrados.append({
                            "nombre": rs,
                            "razon_api": res.get("razon_api", ""),
                            "similitud": res.get("similitud", 0),
                            "resultado": res["resultado"],
                        })

                    filas_procesadas += 1
                    if filas_procesadas % 10 == 0:
                        with _jobs_lock:
                            _jobs[job_id]["procesadas"] = filas_procesadas
                            _jobs[job_id]["progress"] = round(filas_procesadas / total_filas * 100, 1)

            xlsx_bytes = _df_to_xlsx(df)
            with _jobs_lock:
                _jobs[job_id]["archivos"][i].update({
                    "stats": stats,
                    "result": xlsx_bytes,
                    "no_encontrados": no_encontrados,
                    "done": True,
                })

        with _jobs_lock:
            _jobs[job_id].update({
                "status": "done",
                "progress": 100.0,
                "procesadas": total_filas,
                "finished_at": time.time(),
            })

    except Exception as e:
        with _jobs_lock:
            _jobs[job_id].update({"status": "error", "error": str(e)})


# ── Endpoints ─────────────────────────────────────────────────

@router.post("/preview")
async def preview(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Solo .xlsx o .xls")
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents), dtype=str, nrows=5)
        col_empresa = next(
            (c for c in df.columns if re.search(r"raz[oó]n|social|empresa|nombre|cliente|anunciante", str(c), re.I)), ""
        )
        col_ruc = next(
            (c for c in df.columns if re.search(r"\bruc\b|r\.u\.c|registro|contribuyente", str(c), re.I)), ""
        )
        return JSONResponse({
            "columnas": list(df.columns),
            "preview": df.fillna("").to_dict(orient="records"),
            "filename": file.filename,
            "col_empresa_sugerida": col_empresa,
            "col_ruc_sugerida": col_ruc,
        })
    except Exception as e:
        raise HTTPException(500, str(e))


@router.post("/enriquecer")
async def enriquecer(
    files: List[UploadFile] = File(...),
    fuentes: str = Form(...),
    cols_razon: str = Form(...),
    cols_ruc: str = Form("[]"),
):
    fuentes_list = json.loads(fuentes)
    cols_razon_list = json.loads(cols_razon)
    cols_ruc_list = json.loads(cols_ruc)

    if not files:
        raise HTTPException(400, "No se recibieron archivos")

    archivos_data = []
    total_filas = 0

    for i, f in enumerate(files):
        if not f.filename.endswith((".xlsx", ".xls")):
            raise HTTPException(400, f"{f.filename}: formato no soportado")
        try:
            contents = await f.read()
            df = pd.read_excel(io.BytesIO(contents), dtype=str)
        except Exception as e:
            raise HTTPException(400, f"{f.filename}: {e}")

        fuente = fuentes_list[i] if i < len(fuentes_list) else ""
        col_rs_nombre = cols_razon_list[i] if i < len(cols_razon_list) else ""
        col_ruc_nombre = cols_ruc_list[i] if i < len(cols_ruc_list) else ""

        archivos_data.append({
            "nombre": f.filename,
            "fuente_tipo": fuente,
            "df": df,
            "col_razon": _encontrar_col(df, col_rs_nombre) if col_rs_nombre else None,
            "col_ruc": _encontrar_col(df, col_ruc_nombre) if col_ruc_nombre else None,
        })
        total_filas += len(df)

    job_id = str(uuid.uuid4())
    with _jobs_lock:
        _jobs[job_id] = {
            "status": "processing",
            "progress": 0.0,
            "procesadas": 0,
            "total": total_filas,
            "error": None,
            "created_at": time.time(),
            "archivos": [
                {
                    "nombre": a["nombre"],
                    "fuente_tipo": a["fuente_tipo"],
                    "total": len(a["df"]),
                    "done": False,
                    "stats": {},
                    "no_encontrados": [],
                    "result": None,
                }
                for a in archivos_data
            ],
        }

    t = threading.Thread(target=_procesar_job, args=(job_id, archivos_data), daemon=True)
    t.start()

    return JSONResponse({"job_id": job_id, "total": total_filas, "archivos": len(files)})


@router.get("/status/{job_id}")
def status(job_id: str):
    with _jobs_lock:
        job = _jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job no encontrado")
    resp = {
        "status": job["status"],
        "progress": job["progress"],
        "procesadas": job["procesadas"],
        "total": job["total"],
        "archivos": [
            {
                "nombre": a["nombre"],
                "fuente_tipo": a["fuente_tipo"],
                "total": a["total"],
                "done": a["done"],
                "stats": a["stats"],
                "no_encontrados": a["no_encontrados"] if a["done"] else [],
            }
            for a in job["archivos"]
        ],
    }
    if job["status"] == "error":
        resp["error"] = job["error"]
    return JSONResponse(resp)


@router.get("/download/{job_id}")
def download(job_id: str):
    with _jobs_lock:
        job = _jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job no encontrado")
    if job["status"] != "done":
        raise HTTPException(400, f"Job no terminado (status: {job['status']})")

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for a in job["archivos"]:
            if a["result"]:
                nombre_base = a["nombre"].rsplit(".", 1)[0]
                zf.writestr(f"{nombre_base}_ruc.xlsx", a["result"])
    zip_buf.seek(0)

    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=archivos_ruc.zip"},
    )


@router.delete("/job/{job_id}")
def limpiar_job(job_id: str):
    with _jobs_lock:
        if job_id in _jobs:
            del _jobs[job_id]
    return JSONResponse({"ok": True})
